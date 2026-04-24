# breedingSync.py
# import_registry -> writes _temp_import.xlsx
# submit_registry -> reads dirty rows directly from workbook via win32com, sends to ArcGIS

import sys, os, json, datetime
import requests

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
LOG_PATH         = os.path.join(BASE_DIR, "breedingSync.log")
TEMP_IMPORT_PATH = os.path.join(BASE_DIR, "_temp_import.xlsx")
TEMP_RESULT_PATH = os.path.join(BASE_DIR, "_temp_submit_result.json")


def log(msg: str):
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    try:
        print(line)
    except Exception:
        pass
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


# ---------- CONSTANTS ----------

SHEET_REGISTRY = "Реестр опытов"
DIRTY_ALIAS    = "Dirty"

URL_PARENT = "https://maps.ekoniva-apk.org/arcgis/rest/services/breeding/breeding/FeatureServer/0"
URL_CHILD  = "https://maps.ekoniva-apk.org/arcgis/rest/services/breeding/breeding/FeatureServer/1"
SORT_FIELD = "created_date"

FIELDS_PARENT = [
    {"n": "country",          "alias": "Страна",                         "type": "TEXT",   "col": 1},
    {"n": "region",           "alias": "Регион",                         "type": "TEXT",   "col": 2},
    {"n": "site",             "alias": "Опытная площадка",               "type": "TEXT",   "col": 3},
    {"n": "devision",         "alias": "Отделение ЦСиПС",                "type": "TEXT",   "col": 4},
    {"n": "crop",             "alias": "Культура",                       "type": "TEXT",   "col": 5},
    {"n": "farm",             "alias": "Хозяйство (подразделение ПХ)",   "type": "TEXT",   "col": 6},
    {"n": "responsable",      "alias": "Отв. Лицо в ПХ",                 "type": "TEXT",   "col": 7},
    {"n": "fieldNumber",      "alias": "Номер поля",                     "type": "TEXT",   "col": 8},
    {"n": "areaHa",           "alias": "Площадь опыта, га",              "type": "NUMBER", "col": 9},
    {"n": "scheme",           "alias": "Схема опыта",                    "type": "TEXT",   "col": 10},
    {"n": "experimentName",   "alias": "Название опыта",                 "type": "TEXT",   "col": 11},
    {"n": "type",             "alias": "Тип опыта",                      "type": "TEXT",   "col": 12},
    {"n": "productPurpose",   "alias": "Назначение продукции опыта",     "type": "TEXT",   "col": 13},
    {"n": "trialPurpose",     "alias": "Цель, задача опыта",             "type": "TEXT",   "col": 14},
    {"n": "status",           "alias": "Статус опыта",                   "type": "TEXT",   "col": 15},
    {"n": "plantingDate",     "alias": "Дата посева",                    "type": "DATE",   "col": 16},
    {"n": "haverstDate",      "alias": "Дата уборки",                    "type": "DATE",   "col": 17},
    {"n": "report",           "alias": "Отчёт (Выводы, рекомендации)",   "type": "TEXT",   "col": 18},
    {"n": "created_date",     "alias": "created_date",                   "type": "DATE",   "col": 27},
    {"n": "last_edited_date", "alias": "last_edited_date",               "type": "DATE",   "col": 28},
]

CUSTOMER_COLS  = [20, 21, 22, 23]
CUSTOMER_FIELD = "customer"
CUSTOMER_ALIAS = "Заказчик опыта"

DIRTY_COL      = 29
PARENT_GID_COL = 30
CHILD_GID_COL  = 31
TOTAL_COLS     = CHILD_GID_COL

EDITABLE_COLS    = set(range(3, 19))
SYS_SKIP         = {"created_user", "created_date", "last_edited_user", "last_edited_date"}
DATE_ONLY_FIELDS = {"plantingDate", "haverstDate"}

ALIAS_TO_NAME = {f["alias"]: f["n"] for f in FIELDS_PARENT}
ALIAS_TO_NAME[CUSTOMER_ALIAS] = CUSTOMER_FIELD

# ---------- AUTH ----------

TOKEN_URL        = "https://maps.ekoniva-apk.org/portal/sharing/rest/generateToken"
ARC_USERNAME_ENV = "ARCGIS_BREEDING_USER"
ARC_PASSWORD_ENV = "ARCGIS_BREEDING_PASS"


def get_token() -> str:
    username = os.environ.get(ARC_USERNAME_ENV)
    password = os.environ.get(ARC_PASSWORD_ENV)
    if not username or not password:
        raise RuntimeError(f"Не заданы {ARC_USERNAME_ENV}/{ARC_PASSWORD_ENV}")
    payload = {
        "username": username, "password": password,
        "client": "referer", "referer": "https://maps.ekoniva-apk.org",
        "expiration": 60, "f": "json",
    }
    resp = requests.post(TOKEN_URL, data=payload, timeout=30)
    js = resp.json()
    tok = js.get("token")
    if not tok:
        raise RuntimeError(f"Token error: {js}")
    return tok


# ---------- DATE HELPERS ----------

EPOCH       = datetime.datetime(1970, 1, 1)
OFFSET      = datetime.timedelta(hours=3)
EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def esri_ms_to_dt(ms):
    return EPOCH + datetime.timedelta(milliseconds=int(ms)) + OFFSET

def dt_to_esri(dt):
    if dt.tzinfo is not None:
        dt = dt.astimezone(datetime.timezone.utc).replace(tzinfo=None)
    else:
        dt = dt - OFFSET
    return int((dt - EPOCH).total_seconds() * 1000)

def dt_to_excel_serial(dt):
    if dt.tzinfo is not None:
        dt = dt.astimezone(datetime.timezone.utc).replace(tzinfo=None)
    delta = dt - EXCEL_EPOCH
    return delta.days + (delta.seconds + delta.microseconds / 1_000_000) / 86400.0

def excel_serial_to_dt(x):
    return EXCEL_EPOCH + datetime.timedelta(days=float(x))

def arc_value_to_dt(v):
    return esri_ms_to_dt(int(float(v)))


# ---------- QUERY LAYER ----------

def query_layer(url, where="1=1", order_by=""):
    token = get_token()
    session = requests.Session()
    feats, offset, page_size = [], 0, 2000
    while True:
        params = {"where": where, "outFields": "*", "f": "json",
                  "token": token, "resultOffset": offset, "resultRecordCount": page_size}
        if order_by:
            params["orderByFields"] = f"{order_by} DESC"
        r = session.get(url + "/query", params=params, timeout=60)
        js = r.json()
        if "error" in js:
            raise RuntimeError(f"ArcGIS error: {js['error']}")
        page = js.get("features", [])
        feats.extend(page)
        log(f"query_layer offset={offset}: got {len(page)}, total {len(feats)}")
        if not js.get("exceededTransferLimit") or not page:
            break
        offset += len(page)
    return feats


# ---------- IMPORT: write temp xlsx ----------

def import_registry():
    log("=== import_registry START ===")

    try:
        import openpyxl
    except ImportError:
        log("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return 1

    parent_feats = query_layer(URL_PARENT, "1=1", SORT_FIELD)
    log(f"Parent: {len(parent_feats)} features")
    parent_feats.sort(key=lambda f: f.get("attributes", {}).get(SORT_FIELD) or 0, reverse=True)

    child_feats = query_layer(URL_CHILD, "1=1", "")
    log(f"Child: {len(child_feats)} records")

    child_index = {}
    for cf in child_feats:
        attrs = cf.get("attributes", {})
        pgid = attrs.get("parentglobalid")
        if pgid:
            child_index.setdefault(pgid, []).append(attrs.get(CUSTOMER_FIELD))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_REGISTRY

    headers = [""] * TOTAL_COLS
    for f in FIELDS_PARENT:
        col = f.get("col")
        if col:
            headers[col - 1] = f["alias"]
    for c in CUSTOMER_COLS:
        headers[c - 1] = CUSTOMER_ALIAS
    headers[DIRTY_COL - 1]      = DIRTY_ALIAS
    headers[PARENT_GID_COL - 1] = "GlobalID"
    headers[CHILD_GID_COL - 1]  = "ChildGlobalID"
    ws.append(headers)

    fmt_dt   = "DD.MM.YYYY HH:MM"
    fmt_date = "DD.MM.YYYY"

    for ft in parent_feats:
        attrs = ft.get("attributes", {})
        row = [""] * TOTAL_COLS

        for f in FIELDS_PARENT:
            col = f.get("col")
            if not col:
                continue
            v = attrs.get(f["n"])
            if v is None:
                continue
            if f["type"] == "DATE" and isinstance(v, (int, float)):
                row[col - 1] = arc_value_to_dt(v)
            else:
                row[col - 1] = v

        parent_gid = attrs.get("GlobalID", "")
        customers  = child_index.get(parent_gid, [])
        child_gid  = ""
        if customers:
            for cf in child_feats:
                if cf.get("attributes", {}).get("parentglobalid") == parent_gid:
                    child_gid = cf.get("attributes", {}).get("GlobalID", "")
                    break
        for i, c in enumerate(CUSTOMER_COLS):
            if i < len(customers):
                row[c - 1] = customers[i] if customers[i] is not None else ""

        row[DIRTY_COL - 1]      = False
        row[PARENT_GID_COL - 1] = parent_gid
        row[CHILD_GID_COL - 1]  = child_gid
        ws.append(row)

    for f in FIELDS_PARENT:
        col = f.get("col")
        if not col or f["type"] != "DATE":
            continue
        fmt = fmt_date if f["n"] in DATE_ONLY_FIELDS else fmt_dt
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, datetime.datetime):
                cell.number_format = fmt

    if os.path.exists(TEMP_IMPORT_PATH):
        os.remove(TEMP_IMPORT_PATH)
    wb.save(TEMP_IMPORT_PATH)
    log(f"import_registry complete: {ws.max_row - 1} rows -> {TEMP_IMPORT_PATH}")
    return 0


# ---------- SUBMIT: read dirty rows via win32com, send to ArcGIS ----------

def submit_registry(wb_path: str):
    log("=== submit_registry START ===")

    try:
        import win32com.client as win32
    except ImportError:
        log("ERROR: pywin32 not installed. Run: pip install pywin32")
        return 1

    # attach to already-open workbook or open it
    xl = win32.Dispatch("Excel.Application")
    abs_path = os.path.abspath(wb_path)
    wb = None
    opened_here = False
    for book in xl.Workbooks:
        if os.path.abspath(book.FullName) == abs_path:
            wb = book
            break
    if wb is None:
        wb = xl.Workbooks.Open(abs_path)
        opened_here = True

    try:
        try:
            sh = wb.Worksheets(SHEET_REGISTRY)
        except Exception:
            log(f"ERROR: sheet '{SHEET_REGISTRY}' not found in {wb_path}")
            return 1

        last_col = sh.Cells(1, sh.Columns.Count).End(-4159).Column  # xlToLeft
        last_row = sh.Cells(sh.Rows.Count, 1).End(-4162).Row        # xlUp

        if last_row < 2:
            log("No data rows")
            _write_result([])
            return 0

        hdr_vals = list(sh.Range(sh.Cells(1, 1), sh.Cells(1, last_col)).Value[0])

        def col_idx(name):
            for i, h in enumerate(hdr_vals):
                if h == name:
                    return i + 1
            return 0

        dirty_col = col_idx(DIRTY_ALIAS)
        gid_col   = col_idx("GlobalID")

        if not dirty_col:
            log("ERROR: 'Dirty' column not found")
            return 1
        if not gid_col:
            log("ERROR: 'GlobalID' column not found")
            return 1

        # extend last_row by dirty/gid columns to avoid short-range cuts
        last_row = max(
            last_row,
            sh.Cells(sh.Rows.Count, dirty_col).End(-4162).Row,
            sh.Cells(sh.Rows.Count, gid_col).End(-4162).Row,
        )

        data = sh.Range(sh.Cells(2, 1), sh.Cells(last_row, last_col)).Value
        if not data:
            log("No data")
            _write_result([])
            return 0

        name_to_type = {f["n"]: f["type"] for f in FIELDS_PARENT}

        token    = get_token()
        edits    = []
        row_map  = []  # (excel_row, dirty_col) for marking success

        for r_idx, row in enumerate(data, start=2):
            row = list(row)
            dirty_val = row[dirty_col - 1]
            if not dirty_val:
                continue

            parent_gid = row[gid_col - 1]
            if not parent_gid:
                log(f"Row {r_idx}: no GlobalID, skip")
                continue

            attrs = {"GlobalID": str(parent_gid).strip()}

            for c_idx, alias in enumerate(hdr_vals, start=1):
                if not alias or alias == DIRTY_ALIAS:
                    continue
                if alias in ("GlobalID", "ChildGlobalID"):
                    continue

                field_name = ALIAS_TO_NAME.get(alias, alias)
                if field_name.lower() in SYS_SKIP:
                    continue
                f_type = name_to_type.get(field_name)
                if f_type is None:
                    continue

                v = row[c_idx - 1]

                if v in ("", None):
                    attrs[field_name] = None
                elif f_type == "DATE":
                    if isinstance(v, datetime.datetime):
                        attrs[field_name] = dt_to_esri(v)
                    elif isinstance(v, (int, float)):
                        attrs[field_name] = dt_to_esri(excel_serial_to_dt(float(v)))
                    else:
                        try:
                            attrs[field_name] = dt_to_esri(excel_serial_to_dt(float(v)))
                        except Exception:
                            attrs[field_name] = None
                elif f_type == "NUMBER":
                    try:
                        attrs[field_name] = float(v)
                    except Exception:
                        attrs[field_name] = None
                else:
                    attrs[field_name] = v

            edits.append({"attributes": attrs})
            row_map.append((r_idx, dirty_col))

        log(f"Dirty rows found: {len(edits)}")

        if not edits:
            log("No dirty rows")
            _write_result([])
            return 0

        log(f"attrs sample: {json.dumps(edits[0]['attributes'], ensure_ascii=False)}")

        feats_json = json.dumps([{"attributes": e["attributes"]} for e in edits])
        resp = requests.post(URL_PARENT + "/applyEdits", data={
            "f": "json", "token": token,
            "rollbackOnFailure": "True", "useGlobalIds": "True",
            "updates": feats_json,
        }, timeout=60)
        js = resp.json()
        log(f"applyEdits response: {json.dumps(js, ensure_ascii=False)[:500]}")

        results = []
        if "error" in js:
            log(f"applyEdits error: {js['error']}")
            results = [{"row": rm[0], "success": False, "error": str(js["error"])}
                       for rm in row_map]
        else:
            for (excel_row, d_col), r in zip(row_map, js.get("updateResults", [])):
                ok = bool(r.get("success"))
                err_msg = r.get("error", {}).get("description", "")
                results.append({"row": excel_row, "success": ok, "error": err_msg})
                if ok:
                    sh.Cells(excel_row, d_col).Value = False
                    log(f"Row {excel_row}: OK")
                else:
                    log(f"Row {excel_row}: FAILED - {err_msg}")

        wb.Save()
        _write_result(results)
        log(f"submit_registry complete -> {TEMP_RESULT_PATH}")
        return 0

    finally:
        if opened_here:
            try:
                wb.Close(SaveChanges=True)
            except Exception:
                pass


def _write_result(results):
    with open(TEMP_RESULT_PATH, "w", encoding="utf-8") as f:
        json.dump({"status": "ok", "results": results}, f, ensure_ascii=False)


# ---------- MAIN ----------

def normalize_action(a):
    a = (a or "").strip()
    if a.lower().startswith("action="):
        a = a.split("=", 1)[1].strip()
    return a.lower()


def main(argv=None):
    if argv is None:
        argv = sys.argv
    if len(argv) < 2:
        log("Usage: breedingSync.py <action> [workbook_path]")
        return 1

    action = normalize_action(argv[1])
    log("=== breedingSync START ===")
    log(f"action={action!r}")
    log(f"python={sys.executable}  cwd={os.getcwd()}")

    if action == "import_registry":
        return import_registry() or 0

    if action == "submit_registry":
        if len(argv) < 3:
            log("ERROR: submit_registry requires workbook_path as 2nd argument")
            return 1
        return submit_registry(argv[2]) or 0

    log(f"Unknown action: {action!r}. Available: import_registry, submit_registry")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
